"""
excel_exporter.py
─────────────────
인보이스 데이터를 출고요청서 엑셀 양식으로 변환합니다.

출력 컬럼 구조 (택배출고_요청.xlsx 양식 기준)
  A: 품목명       B: 수량   C: 매입가(vat포함)
  D: 수령자       E: 연락처  F: 주소   G: 우편번호   H: 배송메모
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ─── 헤더 스타일 (원본 양식 기준) ─────────────────────────────────────────────
# 헤더 배경: 테마 lt2 (#E7E6E6), 폰트: 맑은 고딕 Bold 10pt
HEADER_FILL   = PatternFill("solid", fgColor="FFE7E6E6")   # lt2 theme color
HEADER_FONT   = Font(name="맑은 고딕", bold=True, size=10)
HEADER_FONT_RED = Font(name="맑은 고딕", bold=True, size=10, color="FFC00000")  # 매입가 헤더

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")

BODY_FONT = Font(name="맑은 고딕", size=10)

THIN  = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ─── 컬럼 정의 (헤더명, 열너비) ─────────────────────────────────────────────
COLUMNS = [
    ("A", "품목명",          43.375),
    ("B", "수량",             9.0),
    ("C", "매입가 (vat포함)", 15.625),
    ("D", "수령자",           12.0),
    ("E", "연락처",           14.0),
    ("F", "주소",            43.125),
    ("G", "우편번호",         15.625),
    ("H", "배송메모",         18.0),
]


def export_to_excel(data: dict) -> bytes:
    """
    data 구조:
        {
            "수령자":   str,
            "연락처":   str,
            "주소":     str,
            "우편번호": str,
            "products": [
                {
                    "어드민코드":     str,   # → A열 품목명
                    "한글품목명":     str,   # 어드민코드 없을 경우 fallback
                    "name":           str,   # 영문명 (최후 fallback)
                    "수량" or "quantity":       int,
                    "매입가(VAT포함)" or "unit_price": float,
                },
                ...
            ]
        }

    출력: 품목 1건 = 1행. 수령자/주소 정보는 각 행에 반복 기재.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # ── 열 너비 설정 ──────────────────────────────────────────────────────────
    for col_letter, _, width in COLUMNS:
        ws.column_dimensions[col_letter].width = width

    # ── 헤더 행 (1행) ─────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 20

    for i, (col_letter, header, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.font      = HEADER_FONT_RED if header == "매입가 (vat포함)" else HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER

    # ── 데이터 행 (2행~) ──────────────────────────────────────────────────────
    products = data.get("products", [])
    recipient   = data.get("수령자", "")
    phone       = data.get("연락처", "")
    address     = data.get("주소", "")
    postal_code = data.get("우편번호", "")

    for idx, item in enumerate(products):
        row = idx + 2
        ws.row_dimensions[row].height = 18

        # 품목명: 어드민코드 우선, 없으면 한글품목명, 없으면 영문명
        product_name = (
            item.get("어드민코드")
            or item.get("admin_code")
            or item.get("한글품목명")
            or item.get("korean_name")
            or item.get("영문품목명")
            or item.get("name", "")
        )

        # 수량
        qty_raw = item.get("수량") or item.get("quantity") or 0
        try:
            qty = int(str(qty_raw).replace(",", ""))
        except (ValueError, TypeError):
            qty = 0

        # 매입가
        price_raw = (
            item.get("매입가(VAT포함)")
            or item.get("unit_price")
            or 0
        )
        try:
            price = float(str(price_raw).replace(",", ""))
        except (ValueError, TypeError):
            price = 0.0

        row_values = [
            product_name,   # A: 품목명
            qty,            # B: 수량
            price,          # C: 매입가
            recipient,      # D: 수령자
            phone,          # E: 연락처
            address,        # F: 주소
            postal_code,    # G: 우편번호
            "",             # H: 배송메모 (빈 값, 사용자가 직접 입력)
        ]

        alignments = [LEFT, CENTER, RIGHT, CENTER, CENTER, LEFT, CENTER, LEFT]
        num_formats = [None, "#,##0", "#,##0", None, None, None, None, None]

        for col_i, (value, align, nfmt) in enumerate(
            zip(row_values, alignments, num_formats), 1
        ):
            cell = ws.cell(row=row, column=col_i, value=value)
            cell.font      = BODY_FONT
            cell.alignment = align
            cell.border    = THIN_BORDER
            if nfmt:
                cell.number_format = nfmt

    # 빈 데이터일 경우 안내 행
    if not products:
        ws.cell(row=2, column=1, value="(품목 없음)").font = Font(name="맑은 고딕", size=10, italic=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
